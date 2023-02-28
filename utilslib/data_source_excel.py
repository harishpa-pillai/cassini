import platform
from openpyxl import load_workbook


class ExcelDataSource:
    """ Data Source utils class for reading credentials excel sheet data"""
    def __init__(self, file_name):
        self._file_name = file_name
        self._excel_data = []
        workbook = load_workbook(filename=self._file_name)
        sheet1 = workbook.active
        for row in sheet1.iter_rows(min_row=2, values_only=True):
            self._excel_data.append(row)

    def get_excel_data(self):
        """Returns data from excel sheet as a list"""
        return self._excel_data


def get_credential_data():
    """Function for reading and providing credential details to the test, for parameterization"""

    file_seperator = None
    if platform.system() == "Windows":
        file_seperator = "\\"
    elif platform.system() == "Linux":
        file_seperator = "/"
    file_name = "Credentials.xlsx"
    file_name_with_path = f".{file_seperator}Files{file_seperator}{file_name}"
    excel_sheet = ExcelDataSource(file_name_with_path)
    return excel_sheet.get_excel_data()
